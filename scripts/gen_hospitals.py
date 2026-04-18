from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = REPO_ROOT / "doc" / "Krankenhäuser_D.xlsx"
DEFAULT_OUT_PATH = REPO_ROOT / "lib" / "data" / "hospitals.json"

DISCIPLINES = [
    "notaufnahme",
    "chirurgie",
    "innere",
    "its",
    "neurochir",
    "verbrennung",
    "paediatrie",
    "op",
]

ABTEILUNG_TO_DISCIPLINE: dict[str, list[str]] = {
    "notaufnahme": ["notaufnahme"],
    "rettungsstelle": ["notaufnahme"],
    "zentrale notaufnahme": ["notaufnahme"],
    "allgemeinchirurgie": ["chirurgie"],
    "unfallchirurgie": ["chirurgie"],
    "viszeralchirurgie": ["chirurgie"],
    "gefaesschirurgie": ["chirurgie"],
    "gefaess-chirurgie": ["chirurgie"],
    "thoraxchirurgie": ["chirurgie"],
    "herzchirurgie": ["chirurgie"],
    "orthopaedie": ["chirurgie"],
    "orthopaedische chirurgie": ["chirurgie"],
    "plastische chirurgie": ["chirurgie"],
    "handchirurgie": ["chirurgie"],
    "mund-kiefer-gesichtschirurgie": ["chirurgie"],
    "kopfchirurgie": ["chirurgie"],
    "transplantationschirurgie": ["chirurgie"],
    "neurochirurgie": ["neurochir"],
    "innere medizin": ["innere"],
    "kardiologie": ["innere"],
    "gastroenterologie": ["innere"],
    "pneumologie": ["innere"],
    "nephrologie": ["innere"],
    "onkologie": ["innere"],
    "haematologie": ["innere"],
    "geriatrie": ["innere"],
    "rheumatologie": ["innere"],
    "endokrinologie": ["innere"],
    "infektiologie": ["innere"],
    "immunologie": ["innere"],
    "tropenmedizin": ["innere"],
    "neurologie": ["innere"],
    "diabetologie": ["innere"],
    "intensivstation": ["its"],
    "intensivmedizin": ["its"],
    "interdisziplinaere intensivstation": ["its"],
    "verbrennungsmedizin": ["verbrennung"],
    "schwerbrandverletzte": ["verbrennung"],
    "paediatrie": ["paediatrie"],
    "kinderheilkunde": ["paediatrie"],
    "neonatologie": ["paediatrie"],
    "kinderchirurgie": ["paediatrie", "chirurgie"],
    "kinderkardiologie": ["paediatrie"],
    "op": ["op"],
    "op-saal": ["op"],
    "anaesthesie": ["op"],
    "anaesthesiologie": ["op"],
}

DISCIPLINE_WEIGHT: dict[str, int] = {
    "notaufnahme": 1,
    "chirurgie": 6,
    "innere": 6,
    "neurochir": 2,
    "verbrennung": 2,
    "paediatrie": 3,
}


@dataclass
class Row:
    name: str
    ort: str | None = None
    art: str | None = None
    adresse: str | None = None
    bundesland: str | None = None
    abteilungen: str | None = None
    betten: float | None = None
    intensivbetten: float | None = None
    lat: float | None = None
    lng: float | None = None


def seeded_rng(seed: int):
    state = seed & 0xFFFFFFFF

    def imul32(a: int, b: int) -> int:
        return ((a & 0xFFFFFFFF) * (b & 0xFFFFFFFF)) & 0xFFFFFFFF

    def rng() -> float:
        nonlocal state
        state = (state + 0x6D2B79F5) & 0xFFFFFFFF
        value = state
        value = imul32(value ^ (value >> 15), value | 1)
        value ^= (value + imul32(value ^ (value >> 7), value | 61)) & 0xFFFFFFFF
        value &= 0xFFFFFFFF
        return ((value ^ (value >> 14)) & 0xFFFFFFFF) / 4294967296

    return rng


def hash_str(text: str) -> int:
    value = 2166136261
    for char in text:
        value ^= ord(char)
        value = ((value & 0xFFFFFFFF) * 16777619) & 0xFFFFFFFF
    return value & 0xFFFFFFFF


def clean_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    text = str(value).strip()
    return text or None


def num_cell(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return None
    return None


def parse_address(raw: str | None) -> dict[str, str]:
    if not raw:
        return {"street": "", "plz": "", "city": ""}
    normalized = re.sub(r"\s+", " ", raw.replace("\n", ", ")).strip()
    parts = [part.strip() for part in normalized.split(",")]
    if len(parts) >= 2:
        match = re.match(r"(\d{5})\s+(.+)", parts[-1])
        if match:
            return {
                "street": ", ".join(parts[:-1]),
                "plz": match.group(1),
                "city": match.group(2),
            }
    match = re.search(r"(\d{5})\s+([^,]+)", normalized)
    if match:
        street = normalized[: match.start()].rstrip(", ").strip()
        return {"street": street, "plz": match.group(1), "city": match.group(2)}
    return {"street": normalized, "plz": "", "city": ""}


def normalize_token(token: str) -> str:
    return (
        token.strip()
        .lower()
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )


def extract_disciplines(abteilungen_field: str) -> set[str]:
    out: set[str] = set()
    for raw in abteilungen_field.split(","):
        mapped = ABTEILUNG_TO_DISCIPLINE.get(normalize_token(raw))
        if mapped:
            out.update(mapped)
    return out


def derive_traeger(name: str, art: str | None) -> str:
    lowered = f"{name} {art or ''}".lower()
    if "privat" in lowered:
        return "privat"
    if re.search(
        r"(diakonie|caritas|marien|st\.?\s|sankt|evangelisch|katholisch|barmherzig|franziskus|johanniter|vincentius)",
        lowered,
        re.IGNORECASE,
    ):
        return "freigemeinnuetzig"
    return "oeffentlich"


def derive_stufe(betten: int, disciplines: set[str], art: str | None) -> str:
    is_uni = bool(re.search(r"universit", art or "", re.IGNORECASE))
    has_burn = "verbrennung" in disciplines
    has_neurochir = "neurochir" in disciplines
    has_its = "its" in disciplines
    has_op = "op" in disciplines
    if is_uni or has_burn or betten >= 1200:
        return "maximal"
    if has_neurochir or betten >= 600:
        return "schwerpunkt"
    if (has_its and has_op) or betten >= 250:
        return "regel"
    return "grund"


def build_disciplines(
    present: set[str],
    total_betten: int,
    intensivbetten: float | None,
    rng,
) -> tuple[dict[str, dict[str, Any]], int]:
    mapped: dict[str, dict[str, Any]] = {}

    its_beds = 0
    if "its" in present:
        its_beds = int(intensivbetten) if intensivbetten and intensivbetten > 0 else max(2, round(total_betten * 0.05))

    op_slots = max(2, round(total_betten / 40)) if "op" in present else 0
    other_beds = max(0, total_betten - its_beds)
    weighted = [discipline for discipline in DISCIPLINES if discipline in present and discipline not in {"its", "op"}]
    total_weight = sum(DISCIPLINE_WEIGHT.get(discipline, 1) for discipline in weighted)

    def mk_cap(beds_total: int) -> dict[str, Any]:
        occupancy_ratio = 0.65 + rng() * 0.15
        beds_occupied = round(beds_total * occupancy_ratio)
        surge_capacity = round(beds_total * 0.2)
        staff_on_duty = max(1, round(beds_total * 0.25))
        staff_on_call = max(1, round(beds_total * 0.15))
        return {
            "bedsTotal": beds_total,
            "bedsOccupied": min(beds_occupied, beds_total),
            "surgeCapacity": surge_capacity,
            "surgeActive": False,
            "staffOnDuty": staff_on_duty,
            "staffOnCall": staff_on_call,
        }

    if its_beds > 0:
        mapped["its"] = mk_cap(its_beds)

    distributed = 0
    for discipline in weighted:
        share = (DISCIPLINE_WEIGHT.get(discipline, 1) / total_weight) if total_weight > 0 else 0
        beds = max(5, round(other_beds * share))
        mapped[discipline] = mk_cap(beds)
        distributed += beds

    if "innere" in mapped:
        diff = other_beds - distributed
        if diff != 0:
            mapped["innere"] = mk_cap(max(5, mapped["innere"]["bedsTotal"] + diff))

    return mapped, op_slots


def iter_rows(xlsx_path: Path) -> list[Row]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows: list[Row] = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        name = clean_cell(row[0] if len(row) > 0 else None)
        if not name:
            continue
        rows.append(
            Row(
                name=name,
                ort=clean_cell(row[1] if len(row) > 1 else None),
                art=clean_cell(row[2] if len(row) > 2 else None),
                adresse=clean_cell(row[3] if len(row) > 3 else None),
                bundesland=clean_cell(row[4] if len(row) > 4 else None),
                abteilungen=clean_cell(row[9] if len(row) > 9 else None),
                betten=num_cell(row[11] if len(row) > 11 else None),
                intensivbetten=num_cell(row[12] if len(row) > 12 else None),
                lat=num_cell(row[13] if len(row) > 13 else None),
                lng=num_cell(row[14] if len(row) > 14 else None),
            )
        )
    workbook.close()
    return rows


def generate_payload(xlsx_path: Path) -> dict[str, Any]:
    rows = iter_rows(xlsx_path)
    simulated: list[dict[str, Any]] = []
    context: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=1):
        if row.lat is None or row.lng is None:
            continue
        hospital_id = f"H-DE-{str(index).zfill(5)}"
        coords = [row.lng, row.lat]
        has_full_data = bool(row.abteilungen) and isinstance(row.betten, float) and row.betten > 0

        if not has_full_data:
            context_row = {
                "id": hospital_id,
                "name": row.name,
                "coords": coords,
                "art": row.art,
                "betten": int(row.betten) if row.betten is not None else None,
                "ort": row.ort,
                "bundesland": row.bundesland,
            }
            context.append({key: value for key, value in context_row.items() if value is not None})
            continue

        disciplines = extract_disciplines(row.abteilungen or "")
        has_real_emergency_dept = "notaufnahme" in disciplines
        disciplines.add("notaufnahme")
        rng = seeded_rng(hash_str(hospital_id))
        bed_count = int(row.betten or 0)
        discipline_map, op_slots_total = build_disciplines(disciplines, bed_count, row.intensivbetten, rng)
        stufe = derive_stufe(bed_count, disciplines, row.art)
        traeger = derive_traeger(row.name, row.art)
        address = parse_address(row.adresse)
        op_slots = {
            "total": op_slots_total,
            "inUse": round(op_slots_total * (0.3 + rng() * 0.3)) if op_slots_total else 0,
        }
        emergency_beds = max(2, round(bed_count * 0.1)) if has_real_emergency_dept else 0
        simulated.append(
            {
                "id": hospital_id,
                "name": row.name,
                "traeger": traeger,
                "versorgungsstufe": stufe,
                "coords": coords,
                "address": {
                    "street": address["street"],
                    "plz": address["plz"],
                    "city": address["city"] or (row.ort or ""),
                    "bundesland": row.bundesland or "",
                },
                "disciplines": discipline_map,
                "opSlots": op_slots,
                "emergencyBeds": emergency_beds,
                "excludedFromAllocation": False,
                "escalationLevel": "normal",
                "canEscalateTo": "katastrophe",
            }
        )

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "simulated": simulated,
        "context": context,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate lib/data/hospitals.json from the source Excel file.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH, help="Path to the source Excel file.")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT_PATH, help="Path to write the hospitals JSON.")
    args = parser.parse_args()

    print(f"[gen-hospitals.py] reading {args.xlsx}")
    payload = generate_payload(args.xlsx)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(
        f"[gen-hospitals.py] simulated={len(payload['simulated'])} "
        f"context={len(payload['context'])}"
    )
    print(f"[gen-hospitals.py] wrote {args.out}")


if __name__ == "__main__":
    main()
